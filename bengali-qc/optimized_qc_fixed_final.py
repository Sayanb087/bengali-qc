"""
Bengali Transcription QC — Local Web Server
Run: python server.py
Then open: http://localhost:5000

Requirements:
    pip install requests openpyxl reportlab
"""

import json, re, os, io, threading, time, threading, time
from collections import Counter
from datetime import datetime as dt
from http.server import HTTPServer, BaseHTTPRequestHandler
from socketserver import ThreadingMixIn

try:
    import requests as req_lib
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable, PageBreak)
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

# ─────────────────────────────────────────────────────────────────
# SERVER-SIDE RESULT CACHE (keyed by filename, stores fixed_data)
# ─────────────────────────────────────────────────────────────────
_fixed_cache = {}  # filename → fixed_data JSON string

# ─────────────────────────────────────────────────────────────────
# VALIDATION LOGIC
# ─────────────────────────────────────────────────────────────────

REQUIRED_KEYS     = {"start", "end", "speaker", "text"}
TIMESTAMP_PATTERN = re.compile(r"^\d{2}:\d{2}:\d{2}$")
TIMESTAMP_LOOSE   = re.compile(r"^\d{1,2}:\d{2}:\d{2}$")
BENGALI_RANGE     = re.compile(r"[\u0980-\u09FF]")
GDRIVE_FILE_RE    = re.compile(r"drive\.google\.com/file/d/([a-zA-Z0-9_-]+)")
GDRIVE_ID_RE      = re.compile(r"id=([a-zA-Z0-9_-]+)")

def extract_drive_id(url):
    m = GDRIVE_FILE_RE.search(url)
    if m: return m.group(1)
    m = GDRIVE_ID_RE.search(url)
    if m: return m.group(1)
    return None

def is_drive_url(s):
    return s.startswith("http") and "drive.google.com" in s

def download_from_drive(url):
    if not HAS_REQUESTS:
        return None, None, "requests library not installed. Run: pip install requests"
    file_id = extract_drive_id(url)
    if not file_id:
        return None, None, "Could not extract file ID from URL. Make sure you copied the full Google Drive link."
    download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
    try:
        session  = req_lib.Session()
        response = session.get(download_url, timeout=30)
        if "virus scan warning" in response.text.lower():
            token = re.search(r'confirm=([0-9A-Za-z_\-]+)', response.text)
            if token:
                download_url = f"https://drive.google.com/uc?export=download&confirm={token.group(1)}&id={file_id}"
                response = session.get(download_url, timeout=30)
        if response.status_code == 200:
            content = response.text.strip()
            if content.startswith("<"):
                return None, None, "Google Drive returned an HTML page. Make sure the file is shared as 'Anyone with the link can view'."
            cd = response.headers.get("Content-Disposition", "")
            fm = re.search(r'filename\*?=["\']?(?:UTF-8\'\')?([^"\';\n]+)', cd)
            filename = fm.group(1).strip() if fm else f"drive_{file_id[:10]}.json"
            return content, filename, None
        elif response.status_code == 403:
            return None, None, "Access denied (403). Share the file as 'Anyone with the link can view' in Google Drive."
        elif response.status_code == 404:
            return None, None, "File not found (404). Check the link is correct."
        else:
            return None, None, f"HTTP {response.status_code} from Google Drive."
    except Exception as ex:
        return None, None, str(ex)

def ts_to_seconds(ts):
    h, m, s = map(int, ts.split(":"))
    return h * 3600 + m * 60 + s

def is_valid_ts(ts):
    return isinstance(ts, str) and bool(TIMESTAMP_PATTERN.match(ts))

def validate(data):
    issues = []
    if not isinstance(data, list):
        issues.append({"level":"ERROR","category":"STRUCTURAL","segment":None,"field":"root",
            "problem":"Top-level is not a JSON array [ ].",
            "fix":"Wrap all segments inside square brackets: [ { ... }, { ... } ]"})
        return issues
    if len(data) == 0:
        issues.append({"level":"WARNING","category":"STRUCTURAL","segment":None,"field":"root",
            "problem":"Array is empty — no segments found.",
            "fix":"Check if the content was accidentally deleted."})
        return issues

    seen_full = {}
    speaker_labels = []

    for i, seg in enumerate(data):
        if not isinstance(seg, dict):
            issues.append({"level":"ERROR","category":"STRUCTURAL","segment":i,"field":"segment",
                "problem":f"Segment is a {type(seg).__name__}, not a JSON object.",
                "fix":"Each segment must be wrapped in curly braces { }."})
            continue
        actual_keys = set(seg.keys())

        for key in REQUIRED_KEYS:
            if key not in actual_keys:
                wrong = next((k for k in actual_keys if k.lower() == key), None)
                if wrong:
                    issues.append({"level":"ERROR","category":"STRUCTURAL","segment":i,"field":key,
                        "problem":f"Key written as '{wrong}' instead of '{key}'.",
                        "fix":f"Rename '{wrong}' to '{key}' (all lowercase)."})
                else:
                    issues.append({"level":"ERROR","category":"STRUCTURAL","segment":i,"field":key,
                        "problem":f"Required key '{key}' is missing.",
                        "fix":f"Add \"{key}\": \"...\" to this segment."})

        known = REQUIRED_KEYS | {k for k in actual_keys if k.lower() in REQUIRED_KEYS}
        for k in actual_keys - known:
            issues.append({"level":"ERROR","category":"STRUCTURAL","segment":i,"field":k,
                "problem":f"Unexpected key '{k}'. Only start, end, speaker, text are allowed.",
                "fix":f"Remove the '{k}' field from this segment."})

        for key in REQUIRED_KEYS:
            if key in actual_keys and seg.get(key) is None:
                issues.append({"level":"ERROR","category":"STRUCTURAL","segment":i,"field":key,
                    "problem":f"'{key}' is set to null.",
                    "fix":f"Replace null with a proper value for '{key}'."})

        for ts_field in ("start","end"):
            val = seg.get(ts_field)
            if val is None: continue
            if isinstance(val, (int, float)):
                issues.append({"level":"ERROR","category":"STRUCTURAL","segment":i,"field":ts_field,
                    "problem":f"'{ts_field}' is a number ({val}), not a string.",
                    "fix":f"Change to a quoted string e.g. \"00:00:{int(val):02d}\"."})
                continue
            if not isinstance(val, str):
                issues.append({"level":"ERROR","category":"STRUCTURAL","segment":i,"field":ts_field,
                    "problem":f"'{ts_field}' must be a string, got {type(val).__name__}.",
                    "fix":"Wrap the value in double quotes."})
                continue
            if not TIMESTAMP_PATTERN.match(val):
                if TIMESTAMP_LOOSE.match(val):
                    fixed = ":".join(p.zfill(2) for p in val.split(":"))
                    issues.append({"level":"ERROR","category":"STRUCTURAL","segment":i,"field":ts_field,
                        "problem":f"'{ts_field}' \"{val}\" is missing zero-padding.",
                        "fix":f"Change \"{val}\" to \"{fixed}\"."})
                else:
                    issues.append({"level":"ERROR","category":"STRUCTURAL","segment":i,"field":ts_field,
                        "problem":f"'{ts_field}' \"{val}\" is not a valid timestamp.",
                        "fix":"Use format HH:MM:SS e.g. \"00:01:35\"."})
                continue
            mm, ss = int(val.split(":")[1]), int(val.split(":")[2])
            if mm > 59 or ss > 59:
                issues.append({"level":"ERROR","category":"LOGICAL","segment":i,"field":ts_field,
                    "problem":f"'{ts_field}' \"{val}\" has impossible values (MM/SS must be 0–59).",
                    "fix":"Correct the timestamp — minutes and seconds cannot exceed 59."})

        s, e = seg.get("start"), seg.get("end")
        if is_valid_ts(s) and is_valid_ts(e):
            ss2, es = ts_to_seconds(s), ts_to_seconds(e)
            if ss2 > es:
                issues.append({"level":"ERROR","category":"LOGICAL","segment":i,"field":"start/end",
                    "problem":f"start \"{s}\" is after end \"{e}\" — timestamps are reversed.",
                    "fix":"Swap the start and end values, or correct whichever is wrong."})
            elif (es - ss2) < 1:
                issues.append({"level":"ERROR","category":"LOGICAL","segment":i,"field":"start/end",
                    "problem":f"Segment duration less than 1 second (start \"{s}\" → end \"{e}\"). Minimum is 1 second.",
                    "fix":"Adjust start or end so there is at least 1 full second. e.g. 00:00:17 → 00:00:18"})

        speaker = seg.get("speaker")
        if speaker is not None:
            if not isinstance(speaker, str):
                issues.append({"level":"ERROR","category":"STRUCTURAL","segment":i,"field":"speaker",
                    "problem":f"'speaker' is a {type(speaker).__name__}, not a string.",
                    "fix":"Wrap the value in double quotes e.g. \"Speaker 1\"."})
            elif speaker.strip() == "":
                issues.append({"level":"ERROR","category":"CONTENT","segment":i,"field":"speaker",
                    "problem":"'speaker' is empty.",
                    "fix":"Fill in the speaker label e.g. \"Speaker 1\"."})
            else:
                speaker_labels.append((i, speaker))

        text = seg.get("text")
        if text is not None:
            if not isinstance(text, str):
                issues.append({"level":"ERROR","category":"STRUCTURAL","segment":i,"field":"text",
                    "problem":f"'text' is a {type(text).__name__}, not a string.",
                    "fix":"Wrap the text value in double quotes."})
            else:
                if text.strip() == "":
                    issues.append({"level":"ERROR","category":"CONTENT","segment":i,"field":"text",
                        "problem":"'text' is empty or contains only whitespace.",
                        "fix":"Fill in the transcription, or remove this segment if added by mistake."})
                else:
                    if "\n" in text or "\r" in text:
                        issues.append({"level":"ERROR","category":"CONTENT","segment":i,"field":"text",
                            "problem":"'text' contains a newline/line break.",
                            "fix":"Remove the line break — text must be a single line."})

                    # ── R: No English letters (a-z, A-Z) ───────────────
                    import re as _re2
                    eng_letters = _re2.findall(r'[a-zA-Z]', text)
                    if eng_letters:
                        unique_letters = sorted(set(eng_letters))
                        # Find the actual words containing English letters for context
                        eng_words = _re2.findall(r'[a-zA-Z][a-zA-Z0-9]*|[a-zA-Z]', text)
                        unique_words = sorted(set(eng_words))
                        issues.append({"level":"ERROR","category":"CONTENT","segment":i,"field":"text",
                            "problem":(f"English letters found in text: {unique_words}. "
                                       f"All text must be written in Bengali script only."),
                            "fix":(f"Replace English letters with Bengali script equivalents. "
                                   f"e.g. \'class\' → \'ক্লাস\', \'COVID\' → \'কোভিড\', "
                                   f"\'USA\' → \'ইউএসএ\'.")})

                    # ── R: No English/Arabic digits (0-9) ───────────────
                    eng_digits = _re2.findall(r'[0-9]', text)
                    if eng_digits:
                        # Find digit sequences for better context
                        digit_seqs = _re2.findall(r'[0-9]+', text)
                        unique_seqs = sorted(set(digit_seqs))
                        issues.append({"level":"ERROR","category":"CONTENT","segment":i,"field":"text",
                            "problem":(f"English/Arabic digits found in text: {unique_seqs}. "
                                       f"All numbers must be written in Bengali digits only."),
                            "fix":(f"Replace English digits with Bengali digits. "
                                   f"e.g. \'500\' → \'৫০০\', \'10000\' → \'১০০০০\', "
                                   f"\'40\' → \'৪০\'.")})


        sig = (seg.get("start"), seg.get("end"), seg.get("speaker"), seg.get("text"))
        if None not in sig:
            if sig in seen_full:
                issues.append({"level":"ERROR","category":"LOGICAL","segment":i,"field":"segment",
                    "problem":f"Exact duplicate of Segment [{seen_full[sig]}] (same start, end, speaker, text).",
                    "fix":"Delete one — it was likely copy-pasted by mistake."})
            else:
                seen_full[sig] = i

    valid_segs = [(i, data[i].get("start"), data[i].get("end")) for i in range(len(data))
                  if isinstance(data[i], dict)
                  and is_valid_ts(data[i].get("start"))
                  and is_valid_ts(data[i].get("end"))]
    for idx in range(1, len(valid_segs)):
        pi, p_start, p_end = valid_segs[idx - 1]
        ci, c_start, c_end = valid_segs[idx]
        p_end_sec   = ts_to_seconds(p_end)
        c_start_sec = ts_to_seconds(c_start)
        p_start_sec = ts_to_seconds(p_start)
        if c_start_sec < p_start_sec:
            issues.append({"level":"ERROR","category":"LOGICAL","segment":ci,"field":"start",
                "problem":f"Segment [{ci}] start \"{c_start}\" is before Segment [{pi}] start \"{p_start}\" — out of order.",
                "fix":f"Re-order segments chronologically. Segment [{ci}] should come before Segment [{pi}]."})
        elif c_start_sec < p_end_sec:
            issues.append({"level":"ERROR","category":"LOGICAL","segment":ci,"field":"start",
                "problem":f"Segment [{ci}] starts at \"{c_start}\" but Segment [{pi}] doesn't end until \"{p_end}\" — overlap.",
                "fix":f"Change start of Segment [{ci}] to \"{p_end}\" or later, or fix the end of Segment [{pi}]."})

    # ── Speaker alternation check ────────────────────────────────
    # Rule: consecutive segments must not have the same speaker
    # (A → B → A → B is correct; A → A or B → B is a mistake)
    speaker_seq = [(i, data[i].get("speaker")) for i in range(len(data))
                   if isinstance(data[i], dict)
                   and isinstance(data[i].get("speaker"), str)
                   and data[i].get("speaker","").strip() != ""]
    for idx in range(1, len(speaker_seq)):
        prev_i, prev_spk = speaker_seq[idx - 1]
        curr_i, curr_spk = speaker_seq[idx]
        if prev_spk == curr_spk:
            issues.append({"level":"ERROR","category":"LOGICAL","segment":curr_i,"field":"speaker",
                "problem":f"Segment [{curr_i}] has the same speaker \"{curr_spk}\" as the previous "
                          f"Segment [{prev_i}]. Speakers must alternate (A → B → A → B).",
                "fix":f"Check if Segment [{curr_i}] belongs to the other speaker, "
                      f"or if Segment [{prev_i}] and [{curr_i}] should be merged into one segment."})

    if speaker_labels:
        all_labels = list({lbl for _, lbl in speaker_labels})
        reported = set()
        for i, label in speaker_labels:
            for other in all_labels:
                if other == label: continue
                pair = tuple(sorted([label, other]))
                if label.lower().replace(" ","") == other.lower().replace(" ","") and pair not in reported:
                    issues.append({"level":"WARNING","category":"CONTENT","segment":i,"field":"speaker",
                        "problem":f"Speaker label \"{label}\" looks like the same person as \"{other}\".",
                        "fix":"Use one consistent format for speaker labels throughout the file."})
                    reported.add(pair)

    # ── R06 + R07: Only "Speaker A", "Speaker B" etc. format allowed ──
    # The ONLY valid format is "Speaker " followed by a single uppercase letter.
    # e.g. Speaker A, Speaker B, Speaker C ...
    # Everything else is invalid: S1, SA, Speaker_1, Spkr1, speaker a etc.
    import re as _re
    VALID_LABEL = _re.compile(r'^Speaker [AB]$')

    unique_labels = list({lbl for _, lbl in speaker_labels})
    reported_invalid = set()

    for i, label in speaker_labels:
        if not VALID_LABEL.match(label) and label not in reported_invalid:
            reported_invalid.add(label)
            issues.append({"level":"ERROR","category":"CONTENT","segment":i,"field":"speaker",
                "problem":(f"Speaker label \"{label}\" is not in the required format. "
                           f"The only accepted format is \"Speaker A\" or \"Speaker B\"."),
                "fix":(f"Rename \"{label}\" to \"Speaker A\" or \"Speaker B\". "
                       f"Update all segments with this label consistently.")})

    return issues

def repair_json(raw):
    """
    Attempt common JSON repairs. Returns (repaired_text, list_of_repairs_made).
    Does NOT raise — returns None if repair still fails.
    """
    repairs = []
    text = raw

    # 1. Replace smart/curly quotes with straight quotes
    for bad, good in [('\u201c','"'),('\u201d','"'),('\u2018',"'"),('\u2019',"'"),
                      ('\u00e2\u0080\u009c','"'),('\u00e2\u0080\u009d','"')]:
        if bad in text:
            text = text.replace(bad, good)
            repairs.append("Replaced curly/smart quotes with straight quotes.")

    # 2. Trailing commas before } or ]
    cleaned = re.sub(r',\s*([}\]])', r'\1', text)
    if cleaned != text:
        repairs.append("Removed trailing commas before } or ].")
        text = cleaned

    # 3. Missing comma between } and {  (common when QC adds a new segment without comma)
    cleaned = re.sub(r'}\s*\n\s*{', '},\n{', text)
    if cleaned != text:
        repairs.append("Added missing commas between segments.")
        text = cleaned

    # 4. Single quotes used instead of double quotes around keys/values
    #    Only attempt if double-quote parse still fails after above fixes
    try:
        json.loads(text)
        return text, repairs   # already valid after above fixes
    except json.JSONDecodeError:
        pass

    # 5. Try wrapping bare text in [] if it looks like raw objects
    stripped = text.strip()
    if stripped.startswith('{') and not stripped.startswith('['):
        wrapped = '[' + stripped + ']'
        try:
            json.loads(wrapped)
            repairs.append("Wrapped bare object(s) in a top-level array [ ].")
            return wrapped, repairs
        except json.JSONDecodeError:
            pass

    # Could not repair
    try:
        json.loads(text)
        return text, repairs
    except json.JSONDecodeError:
        return None, repairs



# ─────────────────────────────────────────────────────────────────
# AUTO-FIXER  (trivial issues — no human judgment needed)
# ─────────────────────────────────────────────────────────────────

# Invisible / junk unicode characters that should never appear in text
INVISIBLE_CHARS = {
    0x200B: "zero-width space",
    0x200C: "zero-width non-joiner",
    0x200D: "zero-width joiner",
    0xFEFF: "BOM / zero-width no-break space",
    0x00A0: "non-breaking space",
    0x0009: "tab character",
    0x000D: "carriage return",
}

def auto_fix_segments(data):
    """
    Apply ONLY safe, non-transcription fixes to each segment.

    What it fixes:
      1. Wrong-case keys       e.g. "Speaker" → "speaker", "Start" → "start"
      2. Invisible unicode     e.g. zero-width spaces, BOM, non-breaking spaces
                               These are genuinely not part of transcription —
                               they are invisible characters that sneak in from
                               copy-paste and break downstream processing.

    What it NEVER touches:
      - start / end timestamps
      - speaker labels
      - actual text content including spaces, punctuation, newlines
        (these are transcription decisions made by the QC team)
    """
    import copy
    fixed = copy.deepcopy(data)
    fixes = []

    if not isinstance(fixed, list):
        return fixed, fixes

    # Truly invisible characters — not part of any transcription
    # These have zero visual representation and only cause problems
    INVISIBLE_CHARS = {
        0x200B: "zero-width space",
        0x200C: "zero-width non-joiner",
        0x200D: "zero-width joiner",
        0xFEFF: "BOM / zero-width no-break space",
    }

    for i, seg in enumerate(fixed):
        if not isinstance(seg, dict):
            continue

        # Fix 1: wrong-case keys (Speaker → speaker, Start → start etc.)
        for key in ("start", "end", "speaker", "text"):
            if key not in seg:
                wrong = next((k for k in seg if k.lower() == key), None)
                if wrong:
                    seg[key] = seg.pop(wrong)
                    fixes.append({"segment": i, "field": key,
                                  "what": f"Renamed key '{wrong}' → '{key}'"})

        # Fix 2: truly invisible unicode characters in text
        # (zero-width space, BOM etc. — completely invisible, never intentional)
        if "text" in seg and isinstance(seg["text"], str):
            original = seg["text"]
            cleaned  = ""
            removed  = set()
            for ch in original:
                if ord(ch) in INVISIBLE_CHARS:
                    removed.add(INVISIBLE_CHARS[ord(ch)])
                else:
                    cleaned += ch
            if removed:
                seg["text"] = cleaned
                fixes.append({"segment": i, "field": "text",
                              "what": f"Removed invisible chars: {', '.join(sorted(removed))}"})

    # Reorder keys to standard order: start, end, speaker, text
    KEY_ORDER = ["start", "end", "speaker", "text"]
    for seg in fixed:
        if not isinstance(seg, dict):
            continue
        ordered = {k: seg[k] for k in KEY_ORDER if k in seg}
        for k in seg:
            if k not in ordered:
                ordered[k] = seg[k]
        seg.clear()
        seg.update(ordered)

    return fixed, fixes

def check_source(source):
    source = source.strip()
    if not source:
        return {"source": source, "filename": "", "status": "error", "error": "Empty input."}

    # ── Load raw content ─────────────────────────────────────────
    if is_drive_url(source):
        content, filename, err = download_from_drive(source)
        if err:
            return {"source": source, "filename": source, "status": "download_error", "error": err}
        raw = content
    else:
        if not os.path.exists(source):
            return {"source": source, "filename": source, "status": "download_error",
                    "error": f"File not found: {source}"}
        with open(source, "r", encoding="utf-8") as f:
            raw = f.read()
        filename = os.path.basename(source)

    # ── Try parsing ──────────────────────────────────────────────
    json_error_msg  = None
    repair_notes    = []
    data            = None

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        json_error_msg = str(e)
        # Attempt repair
        repaired, repair_notes = repair_json(raw)
        if repaired is not None:
            try:
                data = json.loads(repaired)
            except json.JSONDecodeError:
                data = None

    # ── If still unparseable, return JSON error only ─────────────
    if data is None:
        return {
            "source":       source,
            "filename":     filename,
            "status":       "parse_error",
            "error":        json_error_msg,
            "repair_notes": repair_notes,
            "issues":       [],
            "error_count":  0,
            "warning_count":0,
        }

    # ── Auto-fix trivial issues before validation ───────────────
    fixed_data, auto_fixes = auto_fix_segments(data)

    # ── Run full validation on the fixed data ────────────────────
    issues  = validate(fixed_data)
    errors  = [x for x in issues if x["level"] == "ERROR"]
    total   = len(fixed_data) if isinstance(fixed_data, list) else 0

    # Store in server-side cache (safe for large files)
    fixed_json_str = json.dumps(fixed_data, ensure_ascii=False, indent=2)
    _fixed_cache[filename] = fixed_json_str

    return {
        "source":         source,
        "filename":       filename,
        "status":         "fail" if (errors or json_error_msg) else "pass",
        "segments":       total,
        "issues":         issues,
        "error_count":    len(errors),
        "warning_count":  len([x for x in issues if x["level"] == "WARNING"]),
        "json_error":     json_error_msg,
        "repair_notes":   repair_notes,
        "repaired":       json_error_msg is not None and data is not None,
        "auto_fixes":     auto_fixes,
        "has_fixed":      len(auto_fixes) > 0,
    }

# ─────────────────────────────────────────────────────────────────
# EXCEL REPORT GENERATOR
# ─────────────────────────────────────────────────────────────────

def generate_excel(results):
    wb = openpyxl.Workbook()

    # ── Colour palette ────────────────────────────────────────────
    C_HEADER_BG  = "1E293B"   # dark slate
    C_HEADER_FG  = "F8FAFC"   # near white
    C_PASS_BG    = "DCFCE7"   # green tint
    C_PASS_FG    = "166534"
    C_FAIL_BG    = "FEE2E2"   # red tint
    C_FAIL_FG    = "991B1B"
    C_WARN_BG    = "FEF9C3"   # yellow tint
    C_WARN_FG    = "854D0E"
    C_ERR_BG     = "FEE2E2"
    C_ERR_FG     = "991B1B"
    C_ALT_ROW    = "F8FAFC"
    C_BORDER     = "CBD5E1"
    C_SECTION_BG = "EFF6FF"   # light blue for category headers

    thin = Side(style="thin", color=C_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfont(bold=False, color="000000", size=10):
        return Font(name="Calibri", bold=bold, color=color, size=size)

    def hfill(color):
        return PatternFill("solid", fgColor=color)

    def cell(ws, row, col, value, bold=False, bg=None, fg="000000",
             align="left", wrap=False, size=10):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Calibri", bold=bold, color=fg, size=size)
        if bg:
            c.fill = hfill(bg)
        c.alignment = Alignment(horizontal=align, vertical="top",
                                 wrap_text=wrap)
        c.border = border
        return c

    # ══════════════════════════════════════════════════════════════
    # SHEET 1 — SUMMARY
    # ══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 4
    ws1.column_dimensions["B"].width = 42
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 12
    ws1.column_dimensions["F"].width = 14

    # Title block
    ws1.row_dimensions[1].height = 8
    ws1.row_dimensions[2].height = 32
    ws1.merge_cells("B2:F2")
    t = ws1.cell(row=2, column=2, value="Bengali Transcription QC — Batch Report")
    t.font = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=16)
    t.fill = hfill(C_HEADER_BG)
    t.alignment = Alignment(horizontal="left", vertical="center")

    ws1.row_dimensions[3].height = 18
    ws1.merge_cells("B3:F3")
    ts = ws1.cell(row=3, column=2,
                   value=f"Generated: {dt.now().strftime('%d %b %Y  %H:%M')}   |   "
                         f"Total: {len(results)}   |   "
                         f"Passed: {sum(1 for r in results if r.get('status')=='pass')}   |   "
                         f"Failed: {sum(1 for r in results if r.get('status')!='pass')}")
    ts.font = Font(name="Calibri", color="64748B", size=10)
    ts.fill = hfill("F1F5F9")
    ts.alignment = Alignment(horizontal="left", vertical="center")

    ws1.row_dimensions[4].height = 6

    # Column headers
    ws1.row_dimensions[5].height = 20
    ws1.column_dimensions["G"].width = 14
    for col, label in enumerate(["#","File Name","Result","Segments","Errors","Warnings","Auto-Fixed"], start=2):
        cell(ws1, 5, col, label, bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG,
             align="center", size=10)

    # Data rows
    for ri, r in enumerate(results):
        row = ri + 6
        ws1.row_dimensions[row].height = 18
        bg = C_ALT_ROW if ri % 2 == 0 else "FFFFFF"

        status = r.get("status","")
        if status == "pass":
            res_text, res_bg, res_fg = "✓  PASS", C_PASS_BG, C_PASS_FG
        else:
            res_text, res_bg, res_fg = "✗  FAIL", C_FAIL_BG, C_FAIL_FG

        cell(ws1, row, 2, ri + 1,         bg=bg,     align="center")
        cell(ws1, row, 3, r.get("filename",""), bg=bg)
        cell(ws1, row, 4, res_text, bold=True, bg=res_bg, fg=res_fg, align="center")
        cell(ws1, row, 5, r.get("segments", "—"), bg=bg, align="center")
        errs  = r.get("error_count", 0)
        warns = r.get("warning_count", 0)
        afix  = len(r.get("auto_fixes", []))
        cell(ws1, row, 6, errs,  bg=C_ERR_BG  if errs  else bg, fg=C_ERR_FG  if errs  else "000000", align="center")
        cell(ws1, row, 7, warns, bg=C_WARN_BG if warns else bg, fg=C_WARN_FG if warns else "000000", align="center")
        cell(ws1, row, 8, afix,  bg="DCFCE7"  if afix  else bg, fg="166534"  if afix  else "000000", align="center")

    # ══════════════════════════════════════════════════════════════
    # SHEET 2 — DETAILED ISSUES
    # ══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Detailed Issues")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 4
    ws2.column_dimensions["B"].width = 38
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 14
    ws2.column_dimensions["G"].width = 48
    ws2.column_dimensions["H"].width = 48

    ws2.row_dimensions[1].height = 8
    ws2.row_dimensions[2].height = 32
    ws2.merge_cells("B2:H2")
    t2 = ws2.cell(row=2, column=2, value="Bengali Transcription QC — Detailed Issues")
    t2.font = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=16)
    t2.fill = hfill(C_HEADER_BG)
    t2.alignment = Alignment(horizontal="left", vertical="center")

    ws2.row_dimensions[3].height = 6

    # Column headers
    ws2.row_dimensions[4].height = 20
    for col, label in enumerate(["File Name","Result","Segment","Category","Level","Problem","How to Fix"], start=2):
        cell(ws2, 4, col, label, bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")

    detail_row = 5
    for r in results:
        status  = r.get("status","")
        fname   = r.get("filename","")
        issues  = r.get("issues", [])

        if status == "pass" and not issues:
            # Single green row for passing files
            ws2.row_dimensions[detail_row].height = 18
            res_text = "✓  PASS"
            cell(ws2, detail_row, 2, fname,     bg=C_PASS_BG)
            cell(ws2, detail_row, 3, res_text, bold=True, bg=C_PASS_BG, fg=C_PASS_FG, align="center")
            for col in [4,5,6,7,8]:
                cell(ws2, detail_row, col, "—", bg=C_PASS_BG, fg="94A3B8", align="center")
            detail_row += 1
        elif status in ("download_error","parse_error"):
            ws2.row_dimensions[detail_row].height = 30
            cell(ws2, detail_row, 2, fname,                  bg=C_FAIL_BG)
            cell(ws2, detail_row, 3, "✗  FAIL", bold=True,   bg=C_FAIL_BG, fg=C_FAIL_FG, align="center")
            cell(ws2, detail_row, 4, "—",                    bg=C_FAIL_BG, fg="94A3B8", align="center")
            cell(ws2, detail_row, 5, status.replace("_"," ").title(), bg=C_FAIL_BG, fg=C_FAIL_FG, align="center")
            cell(ws2, detail_row, 6, "ERROR", bold=True,     bg=C_FAIL_BG, fg=C_FAIL_FG, align="center")
            cell(ws2, detail_row, 7, r.get("error",""),      bg=C_FAIL_BG, wrap=True)
            cell(ws2, detail_row, 8, "Fix the file and re-check.", bg=C_FAIL_BG, fg="94A3B8", wrap=True)
            detail_row += 1
        else:
            # One row per issue
            for iss in issues:
                ws2.row_dimensions[detail_row].height = 42
                level = iss.get("level","")
                cat   = iss.get("category","")

                if level == "ERROR":
                    row_bg, row_fg = C_FAIL_BG, C_FAIL_FG
                else:
                    row_bg, row_fg = C_WARN_BG, C_WARN_FG

                seg = iss.get("segment")
                seg_label = f"Seg [{seg}]" if seg is not None else "Global"

                cell(ws2, detail_row, 2, fname,      bg="FFFFFF")
                cell(ws2, detail_row, 3, "✗  FAIL" if status=="fail" else "✓  PASS",
                     bold=True,
                     bg=C_FAIL_BG if status=="fail" else C_PASS_BG,
                     fg=C_FAIL_FG if status=="fail" else C_PASS_FG, align="center")
                cell(ws2, detail_row, 4, seg_label,  bg=row_bg, fg=row_fg, align="center")
                cell(ws2, detail_row, 5, cat,        bg="F8FAFC", align="center")
                cell(ws2, detail_row, 6, level,      bold=True, bg=row_bg, fg=row_fg, align="center")
                cell(ws2, detail_row, 7, iss.get("problem",""), bg="FFFFFF", wrap=True)
                cell(ws2, detail_row, 8, iss.get("fix",""),     bg="F0FDF4", fg="166534", wrap=True)
                detail_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────
# PDF REPORT GENERATOR
# ─────────────────────────────────────────────────────────────────

def generate_pdf(results):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm,
        title="Bengali Transcription QC Report"
    )

    # ── Colours ──────────────────────────────────────────────────
    C_DARK    = colors.HexColor("#1E293B")
    C_ACCENT  = colors.HexColor("#4F8EF7")
    C_PASS    = colors.HexColor("#22C55E")
    C_FAIL    = colors.HexColor("#EF4444")
    C_WARN    = colors.HexColor("#F59E0B")
    C_PASS_BG = colors.HexColor("#DCFCE7")
    C_FAIL_BG = colors.HexColor("#FEE2E2")
    C_WARN_BG = colors.HexColor("#FEF9C3")
    C_GREY    = colors.HexColor("#64748B")
    C_LGREY   = colors.HexColor("#F1F5F9")
    C_BORDER  = colors.HexColor("#CBD5E1")
    C_FIX_BG  = colors.HexColor("#F0FDF4")
    C_FIX_FG  = colors.HexColor("#166534")
    C_WHITE   = colors.white

    # ── Styles ───────────────────────────────────────────────────
    styles = getSampleStyleSheet()

    def ps(name, parent="Normal", **kw):
        return ParagraphStyle(name, parent=styles[parent], **kw)

    S_TITLE   = ps("Title2",   fontSize=20, textColor=C_WHITE,   leading=26, spaceAfter=0)
    S_SUB     = ps("Sub",      fontSize=9,  textColor=C_GREY,    leading=13, spaceAfter=0)
    S_SECHEAD = ps("SecHead",  fontSize=11, textColor=C_DARK,    leading=16,
                   fontName="Helvetica-Bold", spaceBefore=10, spaceAfter=4)
    S_FNAME   = ps("FName",    fontSize=10, textColor=C_DARK,    leading=13,
                   fontName="Helvetica-Bold")
    S_BODY    = ps("Body2",    fontSize=8,  textColor=C_DARK,    leading=11)
    S_FIX     = ps("Fix",      fontSize=8,  textColor=C_FIX_FG,  leading=11)
    S_MUTED   = ps("Muted",    fontSize=8,  textColor=C_GREY,    leading=11)
    S_PASS    = ps("PassLbl",  fontSize=9,  textColor=C_PASS,    fontName="Helvetica-Bold")
    S_FAIL    = ps("FailLbl",  fontSize=9,  textColor=C_FAIL,    fontName="Helvetica-Bold")

    W = A4[0] - 30*mm   # usable width
    story = []

    # ── Cover / Header block ─────────────────────────────────────
    total  = len(results)
    passed = sum(1 for r in results if r.get("status") == "pass")
    failed = total - passed
    ts     = dt.now().strftime("%d %b %Y  %H:%M")

    header_data = [[
        Paragraph("Bengali Transcription<br/>QC Report", S_TITLE),
        Paragraph(
            f"<b>Generated:</b> {ts}<br/>"
            f"<b>Total files:</b> {total} &nbsp;&nbsp; "
            f"<b>Passed:</b> {passed} &nbsp;&nbsp; "
            f"<b>Failed:</b> {failed}",
            ps("HdrMeta", fontSize=9, textColor=colors.HexColor("#94A3B8"), leading=14))
    ]]
    header_tbl = Table(header_data, colWidths=[W*0.55, W*0.45])
    header_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0,0),(-1,-1), C_DARK),
        ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0),(-1,-1), 12),
        ("RIGHTPADDING", (0,0),(-1,-1), 12),
        ("TOPPADDING",   (0,0),(-1,-1), 14),
        ("BOTTOMPADDING",(0,0),(-1,-1), 14),
        ("ROUNDEDCORNERS", [6]),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 8*mm))

    # ── Summary table ────────────────────────────────────────────
    story.append(Paragraph("Summary", S_SECHEAD))
    story.append(Spacer(1, 2*mm))

    sum_header = ["#", "File Name", "Result", "Segments", "Errors", "Warnings"]
    sum_rows   = [sum_header]
    for ri, r in enumerate(results):
        status = r.get("status","")
        res    = "✓  PASS" if status == "pass" else "✗  FAIL"
        segs   = str(r.get("segments","—"))
        errs   = str(r.get("error_count", "—"))
        warns  = str(r.get("warning_count","—"))
        fname  = r.get("filename","")
        if len(fname) > 45:
            fname = "…" + fname[-42:]
        sum_rows.append([str(ri+1), fname, res, segs, errs, warns])

    sum_tbl = Table(sum_rows,
                    colWidths=[8*mm, W*0.48, 22*mm, 18*mm, 16*mm, 18*mm],
                    repeatRows=1)

    sum_style = [
        # Header row
        ("BACKGROUND",   (0,0),(-1,0),  C_DARK),
        ("TEXTCOLOR",    (0,0),(-1,0),  C_WHITE),
        ("FONTNAME",     (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0,0),(-1,0),  8),
        ("ALIGN",        (0,0),(-1,0),  "CENTER"),
        ("TOPPADDING",   (0,0),(-1,0),  6),
        ("BOTTOMPADDING",(0,0),(-1,0),  6),
        # Data rows
        ("FONTSIZE",     (0,1),(-1,-1), 8),
        ("TOPPADDING",   (0,1),(-1,-1), 5),
        ("BOTTOMPADDING",(0,1),(-1,-1), 5),
        ("ALIGN",        (0,1),(0,-1),  "CENTER"),
        ("ALIGN",        (2,1),(5,-1),  "CENTER"),
        ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
        ("GRID",         (0,0),(-1,-1), 0.4, C_BORDER),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[C_WHITE, C_LGREY]),
    ]
    # Colour result cells
    for ri2, r in enumerate(results):
        row  = ri2 + 1
        bg   = C_PASS_BG if r.get("status")=="pass" else C_FAIL_BG
        fg   = C_PASS    if r.get("status")=="pass" else C_FAIL
        sum_style += [
            ("BACKGROUND", (2,row),(2,row), bg),
            ("TEXTCOLOR",  (2,row),(2,row), fg),
            ("FONTNAME",   (2,row),(2,row), "Helvetica-Bold"),
        ]
        if r.get("error_count",0):
            sum_style += [("BACKGROUND",(4,row),(4,row), C_FAIL_BG),
                          ("TEXTCOLOR", (4,row),(4,row), C_FAIL)]
        if r.get("warning_count",0):
            sum_style += [("BACKGROUND",(5,row),(5,row), C_WARN_BG),
                          ("TEXTCOLOR", (5,row),(5,row), C_WARN)]

    sum_tbl.setStyle(TableStyle(sum_style))
    story.append(sum_tbl)
    story.append(Spacer(1, 8*mm))

    # ── Detailed issues — one section per file ───────────────────
    story.append(Paragraph("Detailed Issues", S_SECHEAD))

    CAT_LABEL = {
        "STRUCTURAL": "Structural  —  broken JSON format",
        "LOGICAL":    "Logical  —  valid JSON but incorrect data",
        "CONTENT":    "Content  —  text and speaker issues",
    }

    for ri, r in enumerate(results):
        story.append(Spacer(1, 3*mm))
        status = r.get("status","")
        fname  = r.get("filename","")
        issues = r.get("issues",[])
        errors = [x for x in issues if x.get("level")=="ERROR"]

        # File name bar
        res_txt = "✓  PASS" if status=="pass" else "✗  FAIL"
        res_clr = C_PASS   if status=="pass" else C_FAIL
        res_bg  = C_PASS_BG if status=="pass" else C_FAIL_BG

        file_bar = Table(
            [[Paragraph(f"{ri+1}.  {fname}", S_FNAME),
              Paragraph(res_txt, ps("RB", fontSize=9, textColor=res_clr,
                                    fontName="Helvetica-Bold"))]],
            colWidths=[W*0.78, W*0.22])
        file_bar.setStyle(TableStyle([
            ("BACKGROUND",   (0,0),(-1,-1), res_bg),
            ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
            ("LEFTPADDING",  (0,0),(0,-1),  8),
            ("RIGHTPADDING", (-1,0),(-1,-1),8),
            ("TOPPADDING",   (0,0),(-1,-1), 6),
            ("BOTTOMPADDING",(0,0),(-1,-1), 6),
            ("ALIGN",        (1,0),(1,-1),  "RIGHT"),
            ("LINEBELOW",    (0,0),(-1,-1), 1, res_clr),
        ]))
        story.append(file_bar)

        # Download/parse error — completely unreadable
        if status in ("download_error","parse_error"):
            err_tbl = Table(
                [[Paragraph(r.get("error","Unknown error"), S_BODY)]],
                colWidths=[W])
            err_tbl.setStyle(TableStyle([
                ("BACKGROUND",   (0,0),(-1,-1), C_FAIL_BG),
                ("LEFTPADDING",  (0,0),(-1,-1), 8),
                ("TOPPADDING",   (0,0),(-1,-1), 5),
                ("BOTTOMPADDING",(0,0),(-1,-1), 5),
                ("GRID",         (0,0),(-1,-1), 0.4, C_BORDER),
            ]))
            story.append(err_tbl)
            continue

        # JSON was broken but auto-repaired — show amber banner before inner errors
        if r.get("json_error"):
            repair_lines = "  ·  ".join(r.get("repair_notes", []))
            banner_text  = (
                f"<b>⚠  JSON had formatting errors — auto-repaired to reveal inner issues.</b><br/>"
                f"<font color='#92400E'>{r.get('json_error','')}</font><br/>"
                + (f"<font color='#065F46'>Repairs: {repair_lines}</font><br/>" if repair_lines else "")
                + "<font color='#78716C'><i>Fix the original JSON structure. "
                  "Inner errors below are based on the repaired version.</i></font>"
            )
            warn_tbl = Table([[Paragraph(banner_text,
                               ps("WarnB", fontSize=8, textColor=colors.HexColor("#92400E"), leading=12))]],
                             colWidths=[W])
            warn_tbl.setStyle(TableStyle([
                ("BACKGROUND",   (0,0),(-1,-1), colors.HexColor("#FFFBEB")),
                ("LEFTPADDING",  (0,0),(-1,-1), 8),
                ("TOPPADDING",   (0,0),(-1,-1), 6),
                ("BOTTOMPADDING",(0,0),(-1,-1), 6),
                ("LINEBELOW",    (0,0),(-1,-1), 1, colors.HexColor("#F59E0B")),
                ("GRID",         (0,0),(-1,-1), 0.4, colors.HexColor("#FDE68A")),
            ]))
            story.append(warn_tbl)

        # Auto-fix banner
        auto_fixes = r.get("auto_fixes", [])
        if auto_fixes:
            fix_lines = "<br/>".join(
                f"⚡ Seg [{f['segment']}] › {f['field']}: {f['what']}"
                for f in auto_fixes
            )
            fix_tbl = Table(
                [[Paragraph(
                    f"<b>Auto-fixed {len(auto_fixes)} trivial issue(s) — no QC action needed</b><br/>"
                    f"<font color='#065F46'>{fix_lines}</font>",
                    ps("FixB", fontSize=8, textColor=colors.HexColor("#065F46"), leading=12)
                )]],
                colWidths=[W])
            fix_tbl.setStyle(TableStyle([
                ("BACKGROUND",   (0,0),(-1,-1), colors.HexColor("#F0FDF4")),
                ("LEFTPADDING",  (0,0),(-1,-1), 8),
                ("TOPPADDING",   (0,0),(-1,-1), 6),
                ("BOTTOMPADDING",(0,0),(-1,-1), 6),
                ("GRID",         (0,0),(-1,-1), 0.4, colors.HexColor("#BBF7D0")),
            ]))
            story.append(fix_tbl)

        # Pass with no issues
        if status == "pass" and not issues:
            if not auto_fixes:
                ok_tbl = Table(
                    [[Paragraph("No issues found — all checks passed.", S_MUTED)]],
                    colWidths=[W])
                ok_tbl.setStyle(TableStyle([
                    ("BACKGROUND",   (0,0),(-1,-1), C_PASS_BG),
                    ("LEFTPADDING",  (0,0),(-1,-1), 8),
                    ("TOPPADDING",   (0,0),(-1,-1), 5),
                    ("BOTTOMPADDING",(0,0),(-1,-1), 5),
                    ("GRID",         (0,0),(-1,-1), 0.4, C_BORDER),
                ]))
                story.append(ok_tbl)
            continue

        # Group by category
        cats = {"STRUCTURAL":[], "LOGICAL":[], "CONTENT":[]}
        for iss in issues:
            c = iss.get("category","CONTENT")
            if c in cats:
                cats[c].append(iss)

        for cat, cat_issues in cats.items():
            if not cat_issues: continue

            # Category subheader
            cat_hdr = Table(
                [[Paragraph(CAT_LABEL[cat],
                            ps("CatH", fontSize=8, textColor=C_GREY,
                               fontName="Helvetica-Bold"))]],
                colWidths=[W])
            cat_hdr.setStyle(TableStyle([
                ("BACKGROUND",   (0,0),(-1,-1), C_LGREY),
                ("LEFTPADDING",  (0,0),(-1,-1), 8),
                ("TOPPADDING",   (0,0),(-1,-1), 4),
                ("BOTTOMPADDING",(0,0),(-1,-1), 4),
                ("LINEBELOW",    (0,0),(-1,-1), 0.5, C_BORDER),
            ]))
            story.append(cat_hdr)

            # Issue rows  [Seg | Field | Problem | Fix]
            issue_header = ["Segment","Field","Problem","How to Fix"]
            issue_rows   = [issue_header]
            for iss in cat_issues:
                seg   = f"Seg [{iss['segment']}]" if iss.get("segment") is not None else "Global"
                level = iss.get("level","ERROR")
                p_col = C_FAIL if level=="ERROR" else C_WARN
                issue_rows.append([
                    Paragraph(seg,               ps("SC", fontSize=8, textColor=p_col, fontName="Helvetica-Bold")),
                    Paragraph(iss.get("field",""),ps("FC", fontSize=8, textColor=C_ACCENT)),
                    Paragraph(iss.get("problem",""), S_BODY),
                    Paragraph(iss.get("fix",""),     S_FIX),
                ])

            issue_tbl = Table(issue_rows,
                              colWidths=[18*mm, 20*mm, W*0.43, W*0.34],
                              repeatRows=1)
            iss_style = [
                # Header
                ("BACKGROUND",   (0,0),(-1,0), C_DARK),
                ("TEXTCOLOR",    (0,0),(-1,0), C_WHITE),
                ("FONTNAME",     (0,0),(-1,0), "Helvetica-Bold"),
                ("FONTSIZE",     (0,0),(-1,0), 8),
                ("ALIGN",        (0,0),(-1,0), "CENTER"),
                ("TOPPADDING",   (0,0),(-1,0), 5),
                ("BOTTOMPADDING",(0,0),(-1,0), 5),
                # Data
                ("FONTSIZE",     (0,1),(-1,-1), 8),
                ("VALIGN",       (0,0),(-1,-1), "TOP"),
                ("TOPPADDING",   (0,1),(-1,-1), 5),
                ("BOTTOMPADDING",(0,1),(-1,-1), 5),
                ("LEFTPADDING",  (0,0),(-1,-1), 5),
                ("RIGHTPADDING", (0,0),(-1,-1), 5),
                ("GRID",         (0,0),(-1,-1), 0.4, C_BORDER),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[C_WHITE, C_LGREY]),
            ]
            # Fix column light green bg
            for row_i in range(1, len(issue_rows)):
                iss_style.append(("BACKGROUND",(3,row_i),(3,row_i), C_FIX_BG))
            issue_tbl.setStyle(TableStyle(iss_style))
            story.append(issue_tbl)

        story.append(Spacer(1, 2*mm))

    # ── Footer via onFirstPage / onLaterPages ────────────────────
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(C_GREY)
        canvas.drawString(15*mm, 8*mm,
            f"Bengali Transcription QC Report  —  {dt.now().strftime('%d %b %Y')}")
        canvas.drawRightString(A4[0]-15*mm, 8*mm, f"Page {doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
    buf.seek(0)
    return buf.read()

# ─────────────────────────────────────────────────────────────────
# HTML UI
# ─────────────────────────────────────────────────────────────────

HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Bengali Transcription QC</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  :root {
    --bg:      #0f1117; --surface: #181c27; --border: #2a2f3e;
    --accent:  #4f8ef7; --pass:    #22c55e; --fail:   #ef4444;
    --warn:    #f59e0b; --text:    #e2e8f0; --muted:  #64748b;
    --mono:    'IBM Plex Mono', monospace;
    --sans:    'IBM Plex Sans', sans-serif;
  }
  body { background:var(--bg); color:var(--text); font-family:var(--sans); min-height:100vh; }

  header {
    border-bottom:1px solid var(--border); padding:18px 36px;
    display:flex; align-items:center; gap:14px; background:var(--surface);
  }
  header .logo { width:32px;height:32px;background:var(--accent);border-radius:6px;display:grid;place-items:center;font-size:16px; }
  header h1 { font-size:15px;font-weight:600;letter-spacing:.02em; }
  header span { font-size:12px;color:var(--muted);margin-left:4px; }

  .layout { display:grid; grid-template-columns:380px 1fr; min-height:calc(100vh - 61px); }

  .left {
    background:var(--surface); border-right:1px solid var(--border);
    padding:28px 24px; display:flex; flex-direction:column; gap:16px;
  }
  .panel-title { font-size:11px;font-weight:600;letter-spacing:.12em;text-transform:uppercase;color:var(--muted); }

  textarea {
    width:100%;height:240px;background:var(--bg);border:1px solid var(--border);
    border-radius:8px;color:var(--text);font-family:var(--mono);font-size:12px;
    line-height:1.7;padding:14px;resize:vertical;outline:none;transition:border-color .2s;
  }
  textarea:focus { border-color:var(--accent); }
  textarea::placeholder { color:var(--muted); }

  .hint {
    font-size:11px;color:var(--muted);line-height:1.6;padding:10px 12px;
    background:rgba(79,142,247,.06);border:1px solid rgba(79,142,247,.15);border-radius:6px;
  }
  .hint strong { color:var(--accent); }

  button#checkBtn {
    width:100%;padding:13px;background:var(--accent);color:#fff;
    font-family:var(--sans);font-size:14px;font-weight:600;border:none;
    border-radius:8px;cursor:pointer;transition:opacity .2s,transform .1s;
  }
  button#checkBtn:hover { opacity:.88; }
  button#checkBtn:active { transform:scale(.98); }
  button#checkBtn:disabled { opacity:.4;cursor:not-allowed; }

  /* Download button */
  .dl-wrap { display:none; gap:8px; }
  .dl-wrap.visible { display:flex; }
  .dl-btn {
    flex:1; padding:10px 8px; border:1px solid var(--border); border-radius:8px;
    background:var(--bg); color:var(--text); font-family:var(--sans);
    font-size:12px; font-weight:500; cursor:pointer; text-align:center;
    transition:all .2s; display:flex; align-items:center; justify-content:center; gap:6px;
  }
  .dl-btn:hover { background:var(--accent); border-color:var(--accent); color:#fff; }
  .dl-btn .icon { font-size:15px; }

  .stats-row { display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px; }
  .stat-box {
    background:var(--bg);border:1px solid var(--border);
    border-radius:8px;padding:12px;text-align:center;
  }
  .stat-box .num { font-family:var(--mono);font-size:22px;font-weight:600; }
  .stat-box .lbl { font-size:10px;color:var(--muted);margin-top:2px;text-transform:uppercase;letter-spacing:.08em; }
  .stat-box.pass .num { color:var(--pass); }
  .stat-box.fail .num { color:var(--fail); }
  .stat-box.total .num { color:var(--accent); }

  .right { padding:28px 32px;overflow-y:auto;display:flex;flex-direction:column;gap:16px; }
  .empty-state { margin:auto;text-align:center;color:var(--muted); }
  .empty-state .icon { font-size:48px;margin-bottom:16px; }
  .empty-state p { font-size:14px;line-height:1.7; }

  .file-card {
    background:var(--surface);border:1px solid var(--border);
    border-radius:10px;overflow:hidden;animation:slideIn .25s ease;
  }
  @keyframes slideIn { from{opacity:0;transform:translateY(8px)} to{opacity:1;transform:translateY(0)} }

  .file-header {
    display:flex;align-items:center;padding:14px 18px;gap:12px;
    cursor:pointer;user-select:none;border-bottom:1px solid transparent;transition:background .15s;
  }
  .file-header:hover { background:rgba(255,255,255,.03); }
  .file-header.has-issues { border-bottom-color:var(--border); }

  .badge {
    font-family:var(--mono);font-size:11px;font-weight:600;
    padding:3px 9px;border-radius:4px;white-space:nowrap;flex-shrink:0;
  }
  .badge.pass  { background:rgba(34,197,94,.15);color:var(--pass); }
  .badge.fail  { background:rgba(239,68,68,.15);color:var(--fail); }
  .badge.error { background:rgba(239,68,68,.15);color:var(--fail); }

  .file-name { font-family:var(--mono);font-size:13px;flex:1;white-space:nowrap;overflow:hidden;text-overflow:ellipsis; }
  .file-meta { font-size:11px;color:var(--muted);white-space:nowrap; }
  .chevron { color:var(--muted);font-size:12px;transition:transform .2s;flex-shrink:0; }
  .file-header.open .chevron { transform:rotate(90deg); }

  .issues-list { display:none;padding:0 18px 18px; }
  .issues-list.open { display:block; }
  .cat-group { margin-top:16px; }
  .cat-label {
    font-size:10px;font-weight:600;letter-spacing:.12em;text-transform:uppercase;
    color:var(--muted);margin-bottom:8px;padding-bottom:6px;border-bottom:1px solid var(--border);
  }
  .issue-row {
    display:grid;grid-template-columns:90px 80px 1fr;gap:8px;
    padding:10px 0;border-bottom:1px solid rgba(255,255,255,.04);font-size:12px;line-height:1.5;
  }
  .issue-row:last-child { border-bottom:none; }
  .issue-seg { font-family:var(--mono);color:var(--muted);font-size:11px;padding-top:1px; }
  .issue-field { font-family:var(--mono);font-size:11px;color:var(--accent);padding-top:1px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis; }
  .issue-problem { color:var(--text); }
  .issue-fix { color:var(--muted);margin-top:3px;font-size:11px; }
  .issue-fix::before { content:"Fix: ";color:var(--pass);font-weight:600; }
  .err-dot { display:inline-block;width:6px;height:6px;border-radius:50%;margin-right:4px;vertical-align:middle; }
  .err-dot.ERROR   { background:var(--fail); }
  .err-dot.WARNING { background:var(--warn); }
  .error-msg { padding:14px 18px;font-size:12px;color:var(--fail);font-family:var(--mono);line-height:1.6; }

  .progress-bar { height:2px;background:var(--border);border-radius:2px;overflow:hidden; }
  .progress-fill { height:100%;background:var(--accent);transition:width .3s;border-radius:2px; }
  .spinner { display:inline-block;width:14px;height:14px;border:2px solid rgba(255,255,255,.2);border-top-color:#fff;border-radius:50%;animation:spin .6s linear infinite;vertical-align:middle;margin-right:6px; }
  @keyframes spin { to{transform:rotate(360deg)} }

  /* Auto-fix banner */
  .autofix-banner {
    margin: 14px 18px 0;
    padding: 12px 14px;
    background: rgba(34,197,94,.07);
    border: 1px solid rgba(34,197,94,.25);
    border-radius: 7px;
    font-size: 12px;
  }
  .afb-title { color: #22c55e; font-weight: 600; margin-bottom: 6px; }
  .afb-list  { display: flex; flex-wrap: wrap; gap: 5px; margin-bottom: 5px; }
  .afb-item  {
    background: rgba(34,197,94,.1); color: #86efac;
    font-size: 11px; font-family: var(--mono);
    padding: 2px 7px; border-radius: 4px;
  }
  .afb-note  { color: #64748b; font-size: 11px; font-style: italic; }

  /* Per-card fix tag and download button */
  .fix-tag {
    font-size: 10px; font-weight: 600; color: #22c55e;
    background: rgba(34,197,94,.1); padding: 2px 7px; border-radius: 4px;
    white-space: nowrap;
  }
  .dl-fixed-btn {
    font-size: 11px; font-weight: 500;
    color: #e2e8f0; background: rgba(79,142,247,.15);
    border: 1px solid rgba(79,142,247,.3); border-radius: 4px;
    padding: 2px 8px; cursor: pointer; white-space: nowrap;
    transition: background .15s;
  }
  .dl-fixed-btn:hover { background: rgba(79,142,247,.3); }

  /* JSON repair warning banner */
  .json-warn-banner {
    margin: 14px 18px 0;
    padding: 12px 14px;
    background: rgba(245,158,11,.08);
    border: 1px solid rgba(245,158,11,.3);
    border-radius: 7px;
    font-size: 12px;
  }
  .jwb-title   { color: #f59e0b; font-weight: 600; margin-bottom: 5px; }
  .jwb-detail  { color: #94a3b8; font-family: var(--mono); font-size: 11px; margin-bottom: 4px; }
  .jwb-repairs { color: #34d399; font-size: 11px; margin-bottom: 4px; }
  .jwb-note    { color: #64748b; font-size: 11px; font-style: italic; }
</style>
</head>
<body>
<header>
  <div class="logo">✓</div>
  <h1>Bengali Transcription QC <span>Validator</span></h1>
</header>
<div class="layout">
  <div class="left">
    <div class="panel-title">Input</div>
    <textarea id="linksInput" placeholder="Paste Google Drive links or local file paths here — one per line

https://drive.google.com/file/d/ABC.../view?usp=drive_link
https://drive.google.com/file/d/DEF.../view?usp=drive_link
C:\Users\you\files\transcript.json"></textarea>

    <div class="hint">
      <strong>Tip:</strong> Paste one link or file path per line.<br>
      Files must be shared as <strong>"Anyone with the link can view"</strong> on Google Drive.
    </div>

    <div class="progress-bar" id="progressBar" style="display:none">
      <div class="progress-fill" id="progressFill" style="width:0%"></div>
    </div>

    <button id="checkBtn" onclick="runCheck()">Run QC Check</button>

    <div class="dl-wrap" id="dlWrap">
      <button class="dl-btn" onclick="downloadReport('excel')">
        <span class="icon">📊</span> Excel Report
      </button>
      <button class="dl-btn" onclick="downloadReport('pdf')">
        <span class="icon">📄</span> PDF Report
      </button>
    </div>

    <div class="stats-row" id="statsRow" style="display:none">
      <div class="stat-box total"><div class="num" id="statTotal">0</div><div class="lbl">Checked</div></div>
      <div class="stat-box pass"><div class="num" id="statPass">0</div><div class="lbl">Passed</div></div>
      <div class="stat-box fail"><div class="num" id="statFail">0</div><div class="lbl">Failed</div></div>
    </div>
  </div>

  <div class="right" id="results">
    <div class="empty-state">
      <div class="icon">📋</div>
      <p>Paste your Google Drive links<br>on the left and click <strong>Run QC Check</strong>.</p>
    </div>
  </div>
</div>

<script>
let allResults = [];  // store all results for report download

async function runCheck() {
  const raw = document.getElementById('linksInput').value.trim();
  if (!raw) return;
  const sources = raw.split('\n').map(l => l.trim()).filter(l => l && !l.startsWith('#'));
  if (!sources.length) return;

  const btn         = document.getElementById('checkBtn');
  const resultsDiv  = document.getElementById('results');
  const statsRow    = document.getElementById('statsRow');
  const progressBar = document.getElementById('progressBar');
  const progressFill= document.getElementById('progressFill');
  const dlWrap      = document.getElementById('dlWrap');

  btn.disabled = true;
  btn.innerHTML = '<span class="spinner"></span> Checking...';
  resultsDiv.innerHTML = '';
  statsRow.style.display = 'none';
  dlWrap.classList.remove('visible');
  progressBar.style.display = 'block';
  progressFill.style.width = '0%';
  allResults = [];

  let passed = 0, failed = 0;

  for (let i = 0; i < sources.length; i++) {
    const src = sources[i];
    progressFill.style.width = `${Math.round((i / sources.length) * 100)}%`;

    const cardId = `card-${i}`;
    const placeholder = document.createElement('div');
    placeholder.className = 'file-card';
    placeholder.id = cardId;
    placeholder.innerHTML = `
      <div class="file-header">
        <span class="badge" style="background:rgba(100,116,139,.15);color:#64748b">...</span>
        <span class="file-name">${escHtml(src.length > 60 ? src.slice(-55) : src)}</span>
        <span class="file-meta"><span class="spinner"></span></span>
      </div>`;
    resultsDiv.appendChild(placeholder);

    try {
      const resp = await fetch('/check', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({source: src})
      });
      const data = await resp.json();
      renderCard(cardId, data);
      allResults.push(data);
      if (data.status === 'pass') passed++; else failed++;
    } catch(e) {
      const errResult = {source: src, filename: src, status: 'download_error', error: e.message};
      document.getElementById(cardId).innerHTML = `
        <div class="file-header">
          <span class="badge error">✗ FAIL</span>
          <span class="file-name">${escHtml(src)}</span>
        </div>
        <div class="error-msg">Network error: ${escHtml(e.message)}</div>`;
      allResults.push(errResult);
      failed++;
    }

    progressFill.style.width = `${Math.round(((i + 1) / sources.length) * 100)}%`;
  }

  document.getElementById('statTotal').textContent = sources.length;
  document.getElementById('statPass').textContent  = passed;
  document.getElementById('statFail').textContent  = failed;
  statsRow.style.display = 'grid';
  dlWrap.classList.add('visible');

  btn.disabled = false;
  btn.innerHTML = 'Run QC Check';
  setTimeout(() => { progressBar.style.display = 'none'; }, 800);
}

async function downloadReport(fmt) {
  if (!allResults.length) return;
  const btns = document.querySelectorAll('.dl-btn');
  btns.forEach(b => b.disabled = true);
  const clickedBtn = fmt === 'pdf'
    ? btns[1] : btns[0];
  const origLabel = clickedBtn.innerHTML;
  clickedBtn.innerHTML = '<span class="spinner" style="border-top-color:#fff"></span> Generating...';

  const endpoint = fmt === 'pdf' ? '/report/pdf' : '/report/excel';
  const ext      = fmt === 'pdf' ? 'pdf' : 'xlsx';

  try {
    const resp = await fetch(endpoint, {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({results: allResults})
    });

    if (!resp.ok) {
      const err = await resp.text();
      alert('Could not generate report: ' + err);
      return;
    }

    const blob = await resp.blob();
    const url  = URL.createObjectURL(blob);
    const a    = document.createElement('a');
    const ts   = new Date().toISOString().slice(0,16).replace('T','_').replace(':','-');
    a.href     = url;
    a.download = `QC_Report_${ts}.${ext}`;
    a.click();
    URL.revokeObjectURL(url);
  } finally {
    clickedBtn.innerHTML = origLabel;
    btns.forEach(b => b.disabled = false);
  }
}

function renderCard(cardId, data) {
  const card = document.getElementById(cardId);
  if (!card) return;
  const status = data.status;
  const fname  = data.filename || data.source || '—';
  const shortName = fname.length > 55 ? '...' + fname.slice(-52) : fname;

  let badgeHtml, metaHtml;
  const fixCount  = (data.auto_fixes || []).length;
  const fixTag    = fixCount ? `<span class="fix-tag">⚡ ${fixCount} auto-fixed</span>` : '';

  if (status === 'pass' && !data.json_error) {
    badgeHtml = `<span class="badge pass">✓ PASS</span>`;
    metaHtml  = `<span class="file-meta" style="color:var(--pass)">${data.segments} segments</span>`;
  } else if (status === 'download_error' || status === 'parse_error') {
    badgeHtml = `<span class="badge fail">✗ FAIL</span>`;
    metaHtml  = '';
  } else {
    badgeHtml = `<span class="badge fail">✗ FAIL</span>`;
    const repairTag = data.repaired ? '<span style="color:var(--warn);font-size:10px"> · JSON repaired</span>' : '';
    metaHtml  = `<span class="file-meta">${data.error_count} error(s) · ${data.warning_count} warning(s)${repairTag}</span>`;
  }

  const hasDetails = status !== 'pass' || (data.issues && data.issues.length > 0) || fixCount || data.json_error;

  // Build download-fixed button — only show if there are auto_fixes
  const dlFixedBtn = fixCount
    ? `<button class="dl-fixed-btn" title="Download file with auto-corrections applied"
         onclick="event.stopPropagation();downloadFixed('${escHtml(fname)}', ${allResults.indexOf(data)})"
         >⬇ Fixed JSON</button>`
    : '';

  card.innerHTML = `
    <div class="file-header ${hasDetails ? 'has-issues' : ''}" onclick="toggleCard(this)">
      ${badgeHtml}
      <span class="file-name" title="${escHtml(fname)}">${escHtml(shortName)}</span>
      <span style="display:flex;align-items:center;gap:6px">${fixTag}${dlFixedBtn}${metaHtml}</span>
      ${hasDetails ? '<span class="chevron">▶</span>' : ''}
    </div>
    ${buildBody(data)}`;
}

function buildBody(data) {
  const status = data.status;

  // Pure download failure — nothing to show inside
  if (status === 'download_error')
    return `<div class="error-msg">${escHtml(data.error)}</div>`;

  // Completely unparseable — show JSON error only
  if (status === 'parse_error')
    return `<div class="error-msg">
      <div style="margin-bottom:6px">❌ <b>Invalid JSON — file could not be parsed at all.</b></div>
      <div style="color:#94a3b8;font-size:11px">${escHtml(data.error)}</div>
      <div style="margin-top:8px;color:#94a3b8;font-size:11px">
        Fix the JSON structure first, then re-run to see all inner errors.
      </div>
    </div>`;

  let html = '<div class="issues-list">';

  // JSON was broken but auto-repaired — show a warning banner
  if (data.json_error) {
    html += `<div class="json-warn-banner">
      <div class="jwb-title">⚠  JSON had formatting errors — auto-repaired to check contents</div>
      <div class="jwb-detail">${escHtml(data.json_error)}</div>`;
    if (data.repair_notes && data.repair_notes.length) {
      html += `<div class="jwb-repairs">Repairs made: ${data.repair_notes.map(escHtml).join(' · ')}</div>`;
    }
    html += `<div class="jwb-note">Fix the original JSON structure. The inner errors below are based on the repaired version.</div>
    </div>`;
  }

  if (status === 'pass' && (!data.issues || !data.issues.length)) {
    if (!data.json_error && !(data.auto_fixes && data.auto_fixes.length)) return '';
  }

  const cats = {
    STRUCTURAL: { label: 'Structural Errors — broken JSON format',    items: [] },
    LOGICAL:    { label: 'Logical Errors — valid JSON but wrong data', items: [] },
    CONTENT:    { label: 'Content Issues — text and speaker',          items: [] },
  };
  for (const iss of (data.issues || [])) {
    if (cats[iss.category]) cats[iss.category].items.push(iss);
  }

  for (const [, cat] of Object.entries(cats)) {
    if (!cat.items.length) continue;
    html += `<div class="cat-group"><div class="cat-label">${cat.label}</div>`;
    for (const iss of cat.items) {
      const seg = iss.segment != null ? `Seg [${iss.segment}]` : 'Global';
      html += `
      <div class="issue-row">
        <div class="issue-seg"><span class="err-dot ${iss.level}"></span>${escHtml(seg)}</div>
        <div class="issue-field">${escHtml(iss.field)}</div>
        <div class="issue-body">
          <div class="issue-problem">${escHtml(iss.problem)}</div>
          <div class="issue-fix">${escHtml(iss.fix)}</div>
        </div>
      </div>`;
    }
    html += `</div>`;
  }

  // ── Auto-corrected section ────────────────────────────────────
  const autoFixes = data.auto_fixes || [];
  if (autoFixes.length) {
    html += `<div class="cat-group">
      <div class="cat-label" style="color:#34d399;border-color:rgba(52,211,153,.25)">
        ⚡ Auto-Corrected (${autoFixes.length}) — already fixed, baked into ⬇ Fixed JSON
      </div>`;
    for (const fx of autoFixes) {
      const seg = fx.segment != null ? `Seg [${fx.segment}]` : 'Global';
      html += `
      <div class="issue-row">
        <div class="issue-seg"><span class="err-dot" style="background:#34d399"></span>${escHtml(seg)}</div>
        <div class="issue-field" style="color:#34d399">${escHtml(fx.field || '')}</div>
        <div class="issue-body">
          <div class="issue-problem" style="color:#a7f3d0">✓ ${escHtml(fx.what || '')}</div>
        </div>
      </div>`;
    }
    html += `</div>`;
  }

  html += `</div>`;
  return html;
}

async function downloadFixed(fname, idx) {
  // Use filename to retrieve from server cache — no large JSON in HTML attributes
  const btn = event.target;
  const origLabel = btn.innerHTML;
  btn.innerHTML = '⏳';
  btn.disabled  = true;

  try {
    const resp = await fetch('/download/fixed', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({filename: fname})
    });
    if (!resp.ok) { alert('Download failed: ' + await resp.text()); return; }
    const blob = await resp.blob();
    const url  = URL.createObjectURL(blob);
    const a    = document.createElement('a');
    const name = fname.replace(/\.[^.]+$/, '');
    const ext  = fname.includes('.') ? fname.slice(fname.lastIndexOf('.')) : '.json';
    a.href     = url;
    a.download = name + '_fixed' + ext;
    a.click();
    URL.revokeObjectURL(url);
  } finally {
    btn.innerHTML = origLabel;
    btn.disabled  = false;
  }
}

function toggleCard(header) {
  const list = header.nextElementSibling;
  if (!list || !list.classList.contains('issues-list')) return;
  const open = list.classList.toggle('open');
  header.classList.toggle('open', open);
}

function escHtml(s) {
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}
</script>
</body>
</html>"""

# ─────────────────────────────────────────────────────────────────
# HTTP HANDLER
# ─────────────────────────────────────────────────────────────────

class ThreadingHTTPServer(ThreadingMixIn, HTTPServer):
    """Handles each request in a separate thread — supports multiple users simultaneously."""
    daemon_threads = True

class Handler(BaseHTTPRequestHandler):
    def handle_error(self, request, client_address):
        pass  # silently ignore connection resets and aborts — normal browser behaviour
    def log_message(self, fmt, *args):
        pass

    def handle_one_request(self):
        try:
            super().handle_one_request()
        except (ConnectionAbortedError, ConnectionResetError, BrokenPipeError):
            pass  # browser closed connection early — not an error  # silence default access logs

    def do_GET(self):
        try:
            if self.path == "/ping":
                self.send_response(200)
                self.send_header("Content-Type", "text/plain")
                self.end_headers()
                self.wfile.write(b"ok")
                return
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(HTML.encode())
        except ConnectionAbortedError:
            pass

    def do_POST(self):
        try:
            length = int(self.headers.get("Content-Length", 0))
            body   = self.rfile.read(length)

            if self.path == "/check":
                try:
                    payload = json.loads(body)
                    result  = check_source(payload.get("source", ""))
                except Exception as e:
                    result = {"status": "error", "error": str(e), "filename": ""}
                resp = json.dumps(result, ensure_ascii=False).encode()
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", len(resp))
                self.end_headers()
                self.wfile.write(resp)

            elif self.path == "/report/excel":
                if not HAS_OPENPYXL:
                    msg = b"openpyxl not installed. Run: pip install openpyxl"
                    self.send_response(500)
                    self.send_header("Content-Type", "text/plain")
                    self.send_header("Content-Length", len(msg))
                    self.end_headers()
                    self.wfile.write(msg)
                    return
                try:
                    payload = json.loads(body)
                    results = payload.get("results", [])
                    xlsx    = generate_excel(results)
                    fname   = f"QC_Report_{dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    self.send_response(200)
                    self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
                    self.send_header("Content-Length", len(xlsx))
                    self.end_headers()
                    self.wfile.write(xlsx)
                except Exception as e:
                    msg = str(e).encode()
                    self.send_response(500)
                    self.send_header("Content-Type", "text/plain")
                    self.send_header("Content-Length", len(msg))
                    self.end_headers()
                    self.wfile.write(msg)

            elif self.path == "/report/pdf":
                if not HAS_REPORTLAB:
                    msg = b"reportlab not installed. Run: pip install reportlab"
                    self.send_response(500)
                    self.send_header("Content-Type", "text/plain")
                    self.send_header("Content-Length", len(msg))
                    self.end_headers()
                    self.wfile.write(msg)
                    return
                try:
                    payload = json.loads(body)
                    results = payload.get("results", [])
                    pdf     = generate_pdf(results)
                    fname   = f"QC_Report_{dt.now().strftime('%Y%m%d_%H%M')}.pdf"
                    self.send_response(200)
                    self.send_header("Content-Type", "application/pdf")
                    self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
                    self.send_header("Content-Length", len(pdf))
                    self.end_headers()
                    self.wfile.write(pdf)
                except Exception as e:
                    msg = str(e).encode()
                    self.send_response(500)
                    self.send_header("Content-Type", "text/plain")
                    self.send_header("Content-Length", len(msg))
                    self.end_headers()
                    self.wfile.write(msg)

            elif self.path == "/download/fixed":
                try:
                    payload  = json.loads(body)
                    filename = payload.get("filename", "file.json")
                    basename = os.path.basename(filename)

                    # Fetch from server-side cache populated during /check
                    fixed_json_str = _fixed_cache.get(basename)
                    if not fixed_json_str:
                        msg = f"No fixed data for '{basename}'. Please re-run the check first.".encode()
                        self.send_response(404)
                        self.send_header("Content-Type", "text/plain")
                        self.send_header("Content-Length", len(msg))
                        self.end_headers()
                        self.wfile.write(msg)
                        return

                    name, ext = os.path.splitext(basename)
                    out_name  = f"{name}_fixed{ext}"
                    content   = fixed_json_str.encode("utf-8")
                    self.send_response(200)
                    self.send_header("Content-Type", "application/json; charset=utf-8")
                    self.send_header("Content-Disposition", f'attachment; filename="{out_name}"')
                    self.send_header("Content-Length", len(content))
                    self.end_headers()
                    self.wfile.write(content)
                except Exception as e:
                    msg = str(e).encode()
                    self.send_response(500)
                    self.send_header("Content-Type", "text/plain")
                    self.send_header("Content-Length", len(msg))
                    self.end_headers()
                    self.wfile.write(msg)

            else:
                self.send_response(404)
                self.end_headers()

        except ConnectionAbortedError:
            pass


# ─────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import socket, threading, time, urllib.request

    PORT     = int(os.environ.get("PORT", 5000))
    RENDER   = os.environ.get("RENDER")          # set automatically by Render
    APP_URL  = os.environ.get("RENDER_EXTERNAL_URL", "")  # e.g. https://your-app.onrender.com

    server   = ThreadingHTTPServer(("0.0.0.0", PORT), Handler)

    try:
        hostname = socket.gethostname()
        local_ip = socket.gethostbyname(hostname)
    except Exception:
        local_ip = "127.0.0.1"

    print(f"\n  ✓  Bengali Transcription QC Server  (multi-user, threaded)")
    if RENDER:
        print(f"  ➜  Running on Render cloud")
        print(f"  ➜  URL: {APP_URL}")
    else:
        print(f"  ➜  Local  :  http://localhost:{PORT}")
        print(f"  ➜  Network:  http://{local_ip}:{PORT}")
        print(f"  ➜  Share via ngrok for remote access.")
    print(f"  Press Ctrl+C to stop.\n")

    if not HAS_OPENPYXL:
        print("  ⚠  Excel reports disabled. Run:  pip install openpyxl")
    if not HAS_REPORTLAB:
        print("  ⚠  PDF reports disabled.   Run:  pip install reportlab")
    print()

    # ── Keep-alive ping (prevents Render free tier from sleeping) ──
    # Pings the server every 10 minutes during the day
    def keep_alive():
        time.sleep(60)   # wait for server to fully start first
        while True:
            try:
                url = f"{APP_URL}/ping" if APP_URL else f"http://localhost:{PORT}/ping"
                urllib.request.urlopen(url, timeout=5)
            except Exception:
                pass
            time.sleep(600)  # ping every 10 minutes

    if RENDER and APP_URL:
        t = threading.Thread(target=keep_alive, daemon=True)
        t.start()
        print("  ✓  Keep-alive enabled (server will not sleep)")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Server stopped.")
    if not HAS_OPENPYXL:
        print("  ⚠  Excel reports disabled. Run:  pip install openpyxl")
    if not HAS_REPORTLAB:
        print("  ⚠  PDF reports disabled.   Run:  pip install reportlab")
    print()
    def keep_alive():
        """Ping the server every 30 seconds to prevent ngrok tunnel sleeping."""
        while True:
            time.sleep(30)
            try:
                import urllib.request
                urllib.request.urlopen(f"http://localhost:{PORT}/ping", timeout=5)
            except Exception:
                pass

    t = threading.Thread(target=keep_alive, daemon=True)
    t.start()

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Server stopped.")
